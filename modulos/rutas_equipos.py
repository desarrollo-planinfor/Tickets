from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, g, send_file, Response, session
from datetime import datetime, timedelta, date
from extensions import db

from sqlalchemy.orm import joinedload
from models import *
from utils import login_required, requiere_permiso
import os
import json
from sqlalchemy import func
from werkzeug.security import generate_password_hash, check_password_hash


equipos_bp = Blueprint('equipos', __name__, template_folder='templates')

# ==================== EQUIPOS Y MANTENCIONES ====================

@equipos_bp.route('/equipos')
@login_required
@requiere_permiso('ver_inventario')
def lista_equipos():
    """Listado de equipos y mantenciones (Admin y Agentes)"""
    page = request.args.get('page', 1, type=int)
    search_query = request.args.get('q', '').strip()
    area_filter = request.args.get('area', '').strip()
    responsable_filter = request.args.get('responsable', '').strip()
    marca_filter = request.args.get('marca', '').strip()
    tab = request.args.get('tab', 'programados').strip()
    
    query = EquipoMantencion.query
    
    # Separar por tipo de requerimiento
    if tab == 'por_uso':
        query = query.filter(EquipoMantencion.requerimiento == 'Calibración / Verificación')
    else:
        query = query.filter((EquipoMantencion.requerimiento != 'Calibración / Verificación') | (EquipoMantencion.requerimiento == None))
    
    if search_query:
        query = query.filter(EquipoMantencion.nombre.ilike(f'%{search_query}%') | EquipoMantencion.codigo.ilike(f'%{search_query}%'))
    if area_filter:
        query = query.filter(EquipoMantencion.area.ilike(f'%{area_filter}%'))
    if responsable_filter:
        query = query.filter(EquipoMantencion.responsable.ilike(f'%{responsable_filter}%'))
    if marca_filter:
        query = query.filter(EquipoMantencion.marca.ilike(f'%{marca_filter}%'))
        
    # Get distinct values for filters
    areas = [r[0] for r in db.session.query(EquipoMantencion.area).distinct().filter(EquipoMantencion.area != None, EquipoMantencion.area != '').all()]
    responsables = [r[0] for r in db.session.query(EquipoMantencion.responsable).distinct().filter(EquipoMantencion.responsable != None, EquipoMantencion.responsable != '').all()]
    marcas = [r[0] for r in db.session.query(EquipoMantencion.marca).distinct().filter(EquipoMantencion.marca != None, EquipoMantencion.marca != '').all()]
    
    equipos_paginados = query.order_by(EquipoMantencion.id.desc()).paginate(page=page, per_page=15, error_out=False)
    
    # Dashboard: calcular conteos de alertas sobre TODOS los equipos
    todos_equipos = EquipoMantencion.query.all()
    total_equipos = len(todos_equipos)
    total_atrasados = sum(1 for e in todos_equipos if e.estado_alerta == 'Rojo')
    total_proximos = sum(1 for e in todos_equipos if e.estado_alerta == 'Amarillo')
    total_al_dia = sum(1 for e in todos_equipos if e.estado_alerta == 'Verde')
    
    import datetime as _dt
    equipos_alerta_7_dias = []
    for e in todos_equipos:
        if e.proxima_mantencion:
            try:
                fecha_prox = _dt.datetime.strptime(e.proxima_mantencion, '%d/%m/%Y').date()
                dias = (fecha_prox - _dt.date.today()).days
                if 0 <= dias <= 7:
                    equipos_alerta_7_dias.append({'equipo': e, 'dias': dias})
            except Exception:
                pass
    equipos_alerta_7_dias.sort(key=lambda x: x['dias'])
    
    return render_template('equipos/lista.html',  
                           equipos_paginados=equipos_paginados, 
                           search_query=search_query,
                           area_filter=area_filter,
                           responsable_filter=responsable_filter,
                           marca_filter=marca_filter,
                           areas=sorted(areas),
                           responsables=sorted(responsables),
                           marcas=sorted(marcas),
                           total_equipos=total_equipos,
                           total_atrasados=total_atrasados,
                           total_proximos=total_proximos,
                           total_al_dia=total_al_dia,
                           equipos_alerta_7_dias=equipos_alerta_7_dias,
                           current_tab=tab)

def normalizar_fecha(fecha_str):
    """Convierte una fecha de YYYY-MM-DD a DD/MM/YYYY si es necesario.
    Si ya está en DD/MM/YYYY la retorna sin cambios."""
    if not fecha_str:
        return fecha_str
    fecha_str = fecha_str.strip()
    # Formato ISO del browser: YYYY-MM-DD
    if len(fecha_str) == 10 and fecha_str[4] == '-' and fecha_str[7] == '-':
        try:
            yyyy, mm, dd = fecha_str.split('-')
            return f"{dd}/{mm}/{yyyy}"
        except Exception:
            pass
    return fecha_str

@equipos_bp.route('/equipos/nuevo', methods=['GET', 'POST'])
@login_required
@requiere_permiso('ver_inventario')
def nuevo_equipo():
    if request.method == 'POST':
        nombre = request.form.get('nombre')
        if nombre == '__otra__':
            nombre = request.form.get('nuevo_nombre')
            
        marca = request.form.get('marca')
        if marca == '__otra__':
            marca = request.form.get('nueva_marca')
            
        requerimiento = request.form.get('requerimiento')
        if requerimiento == '__otra__':
            requerimiento = request.form.get('nuevo_requerimiento')
            
        equipo = EquipoMantencion(
            codigo=request.form.get('codigo'),
            nombre=nombre,
            marca=marca,
            modelo=request.form.get('modelo'),
            serie=request.form.get('serie'),
            area=request.form.get('area'),
            responsable=request.form.get('responsable'),
            ultima_mantencion=normalizar_fecha(request.form.get('ultima_mantencion')),
            frecuencia_mantencion=request.form.get('frecuencia_mantencion'),
            proxima_mantencion=normalizar_fecha(request.form.get('proxima_mantencion')),
            requerimiento=requerimiento,
            tipo_mantencion=request.form.get('tipo_mantencion'),
            estado=request.form.get('estado'),
            ficha=request.form.get('ficha')
        )
        db.session.add(equipo)
        db.session.flush()
        
        responsable = request.form.get('responsable')
        if responsable and responsable.strip():
            hist_resp = HistorialResponsable(
                equipo_id=equipo.id,
                responsable=responsable.strip(),
                fecha_inicio=datetime.now(),
                fecha_fin=None
            )
            db.session.add(hist_resp)

        db.session.commit()
        flash('Equipo registrado exitosamente.', 'success')
        return redirect(url_for('equipos.lista_equipos'))
        
    nombres = [r[0] for r in db.session.query(EquipoMantencion.nombre).distinct().filter(EquipoMantencion.nombre != None, EquipoMantencion.nombre != '').all()]
    marcas = [r[0] for r in db.session.query(EquipoMantencion.marca).distinct().filter(EquipoMantencion.marca != None, EquipoMantencion.marca != '').all()]
    requerimientos = [r[0] for r in db.session.query(EquipoMantencion.requerimiento).distinct().filter(EquipoMantencion.requerimiento != None, EquipoMantencion.requerimiento != '').all()]
    areas_equipo = AreaEquipo.query.order_by(AreaEquipo.nombre).all()
    
    return render_template('equipos/formulario.html', equipo=None, nombres=sorted(nombres), marcas=sorted(marcas), requerimientos=sorted(requerimientos), areas_equipo=areas_equipo)

@equipos_bp.route('/equipos/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@requiere_permiso('ver_inventario')
def editar_equipo(id):
    equipo = db.session.get(EquipoMantencion, id)
    if not equipo:
        flash('Equipo no encontrado.', 'error')
        return redirect(url_for('equipos.lista_equipos'))
        
    if request.method == 'POST':
        nombre = request.form.get('nombre')
        if nombre == '__otra__':
            nombre = request.form.get('nuevo_nombre')
            
        marca = request.form.get('marca')
        if marca == '__otra__':
            marca = request.form.get('nueva_marca')
            
        requerimiento = request.form.get('requerimiento')
        if requerimiento == '__otra__':
            requerimiento = request.form.get('nuevo_requerimiento')
            
        old_resp = (equipo.responsable or '').strip()
        new_resp = (request.form.get('responsable') or '').strip()
        if old_resp != new_resp:
            active_hist = HistorialResponsable.query.filter_by(equipo_id=equipo.id, fecha_fin=None).first()
            if active_hist:
                active_hist.fecha_fin = datetime.now()
            if new_resp:
                new_hist = HistorialResponsable(
                    equipo_id=equipo.id,
                    responsable=new_resp,
                    fecha_inicio=datetime.now(),
                    fecha_fin=None
                )
                db.session.add(new_hist)

        equipo.codigo = request.form.get('codigo')
        equipo.nombre = nombre
        equipo.marca = marca
        equipo.modelo = request.form.get('modelo')
        equipo.serie = request.form.get('serie')
        equipo.area = request.form.get('area')
        equipo.responsable = new_resp
        equipo.ultima_mantencion = normalizar_fecha(request.form.get('ultima_mantencion'))
        equipo.frecuencia_mantencion = request.form.get('frecuencia_mantencion')
        equipo.proxima_mantencion = normalizar_fecha(request.form.get('proxima_mantencion'))
        equipo.requerimiento = requerimiento
        equipo.tipo_mantencion = request.form.get('tipo_mantencion')
        equipo.estado = request.form.get('estado')
        equipo.ficha = request.form.get('ficha')
        
        db.session.commit()
        flash('Equipo actualizado exitosamente.', 'success')
        return redirect(url_for('equipos.lista_equipos'))
        
    nombres = [r[0] for r in db.session.query(EquipoMantencion.nombre).distinct().filter(EquipoMantencion.nombre != None, EquipoMantencion.nombre != '').all()]
    marcas = [r[0] for r in db.session.query(EquipoMantencion.marca).distinct().filter(EquipoMantencion.marca != None, EquipoMantencion.marca != '').all()]
    requerimientos = [r[0] for r in db.session.query(EquipoMantencion.requerimiento).distinct().filter(EquipoMantencion.requerimiento != None, EquipoMantencion.requerimiento != '').all()]
    areas_equipo = AreaEquipo.query.order_by(AreaEquipo.nombre).all()
    
    return render_template('equipos/formulario.html', equipo=equipo, nombres=sorted(nombres), marcas=sorted(marcas), requerimientos=sorted(requerimientos), areas_equipo=areas_equipo)

@equipos_bp.route('/equipos/inactivar/<int:id>', methods=['POST'])
@login_required
@requiere_permiso('ver_inventario')
def inactivar_equipo(id):
    equipo = db.session.get(EquipoMantencion, id)
    if equipo:
        equipo.estado = 'Inactivo'
        db.session.commit()
        flash('Equipo marcado como Inactivo.', 'success')
    return redirect(url_for('equipos.lista_equipos'))

@equipos_bp.route('/equipos/activar/<int:id>', methods=['POST'])
@login_required
@requiere_permiso('ver_inventario')
def activar_equipo(id):
    equipo = db.session.get(EquipoMantencion, id)
    if equipo:
        equipo.estado = 'Operativo'
        db.session.commit()
        flash('Equipo activado exitosamente.', 'success')
    return redirect(url_for('equipos.lista_equipos'))

@equipos_bp.route('/equipos/<int:id>/pdf')
@login_required
@requiere_permiso('ver_inventario')
def descargar_ficha_pdf(id):
    equipo = db.session.get(EquipoMantencion, id)
    if not equipo:
        flash('Equipo no encontrado.', 'error')
        return redirect(url_for('equipos.lista_equipos'))
    
    # Obtener historial más reciente (si existe)
    historial_reciente = HistorialMantencion.query.filter_by(equipo_id=id).order_by(HistorialMantencion.fecha_realizada.desc()).first()
    
    import tempfile
    import os
    from flask import send_file
    from generador_pdf import crear_ficha_pdf
    
    fd, temp_path = tempfile.mkstemp(suffix='.pdf', prefix=f'Ficha_{equipo.codigo or id}_')
    os.close(fd)
    
    crear_ficha_pdf(equipo, historial_reciente, temp_path)
    
    return send_file(
        temp_path, 
        as_attachment=True, 
        download_name=f"Ficha_Mantencion_{equipo.codigo or id}.pdf",
        mimetype='application/pdf'
    )

@equipos_bp.route('/mantencion/<int:id>/pdf')
@login_required
def descargar_pdf_mantencion(id):
    mantencion = db.session.get(HistorialMantencion, id)
    if not mantencion:
        flash('Mantención no encontrada.', 'error')
        return redirect(url_for('equipos.lista_equipos'))
        
    equipo = db.session.get(EquipoMantencion, mantencion.equipo_id)
    
    import tempfile
    import os
    from flask import send_file
    from generador_pdf import crear_ficha_pdf
    
    fd, temp_path = tempfile.mkstemp(suffix='.pdf', prefix=f'Mantencion_{mantencion.id}_')
    os.close(fd)
    
    crear_ficha_pdf(equipo, mantencion, temp_path)
    
    fecha_str = mantencion.fecha_realizada.strftime('%Y%m%d') if mantencion.fecha_realizada else 'sinfecha'
    
    return send_file(
        temp_path, 
        as_attachment=True, 
        download_name=f"Ficha_Mantencion_{equipo.codigo or equipo.id}_{fecha_str}.pdf",
        mimetype='application/pdf'
    )

@equipos_bp.route('/equipos/descargar-masivo', methods=['POST'])
@login_required
@requiere_permiso('ver_inventario')
def descargar_fichas_masivo():
    equipo_ids = request.form.getlist('equipo_ids')
    if not equipo_ids:
        flash('No se seleccionó ningún equipo para descargar.', 'warning')
        return redirect(url_for('equipos.lista_equipos'))
        
    import tempfile
    import os
    import zipfile
    from flask import send_file
    from generador_pdf import crear_ficha_pdf
    
    # Crear un directorio temporal para el ZIP
    temp_dir = tempfile.mkdtemp()
    zip_path = os.path.join(temp_dir, 'Fichas_Mantencion.zip')
    
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for eid in equipo_ids:
            equipo = db.session.get(EquipoMantencion, int(eid))
            if not equipo:
                continue
                
            historial_reciente = HistorialMantencion.query.filter_by(equipo_id=equipo.id).order_by(HistorialMantencion.fecha_realizada.desc()).first()
            
            pdf_filename = f"Ficha_Mantencion_{equipo.codigo or equipo.id}.pdf"
            pdf_path = os.path.join(temp_dir, pdf_filename)
            
            crear_ficha_pdf(equipo, historial_reciente, pdf_path)
            zipf.write(pdf_path, arcname=pdf_filename)
            
    return send_file(
        zip_path,
        as_attachment=True,
        download_name='Fichas_Mantencion_Planinfor.zip',
        mimetype='application/zip'
    )

@equipos_bp.route('/equipos/<int:id>/historial', methods=['GET', 'POST'])
@login_required
@requiere_permiso('ver_inventario')
def historial_equipo(id):
    """Ver historial de mantenciones de un equipo y registrar nuevas"""
    
    equipo = db.session.get(EquipoMantencion, id)
    if not equipo:
        flash('Equipo no encontrado.', 'error')
        return redirect(url_for('equipos.lista_equipos'))
    
    if request.method == 'POST':
        fecha_str = request.form.get('fecha_realizada', '').strip()
        tecnico = request.form.get('tecnico', '').strip()
        observaciones = request.form.get('observaciones', '').strip()
        tipo = request.form.get('tipo', '').strip()
        
        # Recopilar detalles del PDF
        detalles_adicionales = []
        for i in range(1, 4):
            act = request.form.get(f'actividad_{i}')
            if act:
                detalles_adicionales.append(f"Actividad: {act}")
        
        obs = request.form.get('observacion_tecnica')
        if obs:
            detalles_adicionales.append(f"Observacion: {obs}")
            
        res = request.form.get('resultado_pruebas')
        if res:
            detalles_adicionales.append(f"Resultado: {res}")
        
        if detalles_adicionales:
            texto_detalles = "--- Detalles Adicionales ---\n" + "\n".join(detalles_adicionales)
            if observaciones:
                observaciones = observaciones + "\n\n" + texto_detalles
            else:
                observaciones = texto_detalles
        
        # Parsear fecha
        fecha_realizada = datetime.now()
        if fecha_str:
            try:
                fecha_realizada = datetime.strptime(fecha_str, '%d/%m/%Y')
            except ValueError:
                pass
                
        # Capturar actividades adicionales
        actividades = []
        for key in request.form.keys():
            if key.startswith('actividad_tipo_'):
                idx = key.split('_')[-1]
                tipo_act = request.form.get(key, '').strip()
                valor_act = request.form.get(f'actividad_valor_{idx}', '').strip()
                if tipo_act and valor_act:
                    actividades.append(f"- {tipo_act}: {valor_act}")
        
        if actividades:
            texto_actividades = "\n".join(actividades)
            if observaciones:
                observaciones = observaciones + "\n\nActividades adicionales:\n" + texto_actividades
            else:
                observaciones = "Actividades adicionales:\n" + texto_actividades
        
        registro = HistorialMantencion(
            equipo_id=equipo.id,
            fecha_realizada=fecha_realizada,
            tecnico=tecnico,
            observaciones=observaciones,
            tipo=tipo,
            registrado_por=g.usuario.nombre
        )
        db.session.add(registro)
        
        # Actualizar última mantención del equipo
        equipo.ultima_mantencion = fecha_realizada.strftime('%d/%m/%Y')
        
        # Recalcular próxima mantención según frecuencia
        freq = (equipo.frecuencia_mantencion or '').strip().lower()
        if freq:
            from dateutil.relativedelta import relativedelta
            if 'anual' in freq:
                prox = fecha_realizada + relativedelta(years=1)
            elif 'semestral' in freq:
                prox = fecha_realizada + relativedelta(months=6)
            elif 'trimestral' in freq:
                prox = fecha_realizada + relativedelta(months=3)
            elif 'mensual' in freq:
                prox = fecha_realizada + relativedelta(months=1)
            elif 'bimestral' in freq:
                prox = fecha_realizada + relativedelta(months=2)
            else:
                prox = None
            
            if prox:
                equipo.proxima_mantencion = prox.strftime('%d/%m/%Y')
        
        db.session.commit()
        flash('Mantención registrada exitosamente.', 'success')
        return redirect(url_for('equipos.historial_equipo', id=equipo.id))
    
    historial = HistorialMantencion.query.filter_by(equipo_id=equipo.id).order_by(HistorialMantencion.fecha_realizada.desc()).all()
    historial_responsables = HistorialResponsable.query.filter_by(equipo_id=equipo.id).order_by(HistorialResponsable.fecha_inicio.desc()).all()

    return render_template('equipos/historial.html', equipo=equipo, historial=historial, responsables=historial_responsables)

def cron_alertas_mantenciones():
    """Cron: Enviar alerta solo cuando equipos están a 30, 15 o 1 días de mantención"""
    try:
        with app.app_context():
            import datetime as _dt
            equipos = EquipoMantencion.query.all()
            
            # Solo nos interesan los que están en los umbrales específicos
            proximos = []
            for e in equipos:
                if not e.proxima_mantencion:
                    continue
                try:
                    fecha_prox = _dt.datetime.strptime(e.proxima_mantencion, '%d/%m/%Y').date()
                    dias = (fecha_prox - _dt.date.today()).days
                    if dias in [30, 15, 1]:
                        proximos.append((e, dias))
                except Exception:
                    continue
            
            if not proximos:
                return
            
            # Ordenar por días restantes (más urgente primero)
            proximos.sort(key=lambda x: x[1])
            
            # Construir contenido del correo
            filas = _fila_dato('Equipos por mantener', str(len(proximos)), highlight=True)
            
            detalle = '<tr><td style="padding:20px 30px;">'
            detalle += '<p style="margin:0 0 10px;font-size:14px;font-weight:600;color:#f59e0b;">⚠️ Equipos que requieren mantención:</p>'
            detalle += '<ul style="margin:0;padding-left:20px;font-size:13px;color:#374151;">'
            for e, dias in proximos[:20]:
                urgencia = f'<span style="color:#ef4444;font-weight:600;">en {dias} días</span>'
                responsable = e.responsable if e.responsable else "Sin responsable asignado"
                detalle += f'<li><strong>{e.nombre or e.codigo}</strong> — {urgencia} ({e.proxima_mantencion}) — {responsable}</li>'
            if len(proximos) > 20:
                detalle += f'<li>...y {len(proximos) - 20} más</li>'
            detalle += '</ul></td></tr>'
            
            html_correo = _base_email_html(
                '#f59e0b',
                '&#9888; Mantenciones Próximas',
                f'{len(proximos)} equipo(s) necesitan mantención pronto',
                filas,
                detalle,
                target_url=f'{BASE_URL}/equipos'
            )
            
            notif = Notificacion(
                destinatario='ti.noreply@planinfor.cl',
                asunto=f'[Mantenciones] {len(proximos)} equipo(s) necesitan mantención pronto',
                mensaje=html_correo,
                tipo='email',
                estado='pendiente'
            )
            db.session.add(notif)
            db.session.commit()
            
    except Exception as e:
        print(f'Error en cron_alertas_mantenciones: {e}')

def cron_alertas_licencias():
    """Cron: Alerta de licencias a punto de expirar"""
    try:
        with app.app_context():
            import datetime as _dt
            licencias = Licencia.query.all()
            proximos = []
            hoy = _dt.date.today()
            for l in licencias:
                if not l.fecha_expiracion or l.renovacion_automatica:
                    continue
                dias = (l.fecha_expiracion - hoy).days
                if dias in [30, 15, 7, 1, 0]:
                    proximos.append((l, dias))
            
            if not proximos:
                return
                
            proximos.sort(key=lambda x: x[1])
            filas = _fila_dato('Licencias por expirar', str(len(proximos)), highlight=True)
            
            detalle = '<tr><td style="padding:20px 30px;">'
            detalle += '<p style="margin:0 0 10px;font-size:14px;font-weight:600;color:#ef4444;">🚨 Licencias próximas a expirar o expiradas:</p>'
            detalle += '<ul style="margin:0;padding-left:20px;font-size:13px;color:#374151;">'
            for l, dias in proximos:
                urgencia = f'<span style="color:#ef4444;font-weight:600;">en {dias} días</span>' if dias > 0 else f'<span style="color:#ef4444;font-weight:600;">HOY</span>'
                detalle += f'<li><strong>{l.nombre_servicio}</strong> ({l.tipo}) — expira {urgencia} ({l.fecha_expiracion})</li>'
            detalle += '</ul></td></tr>'
            
            html_correo = _base_email_html(
                '#ef4444',
                '&#128680; Licencias por expirar',
                f'{len(proximos)} licencia(s) expiran pronto',
                filas,
                detalle,
                target_url=f'{BASE_URL}/licencias'
            )
            
            notif = Notificacion(
                destinatario='ti.noreply@planinfor.cl',
                asunto=f'[Licencias] {len(proximos)} licencia(s) expiran pronto',
                mensaje=html_correo,
                tipo='email',
                estado='pendiente'
            )
            db.session.add(notif)
            db.session.commit()
    except Exception as e:
        print(f'Error en cron_alertas_licencias: {e}')


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO

@equipos_bp.route('/exportar/inventario')
@login_required
@requiere_permiso('ver_inventario')
def exportar_inventario():
    equipos = EquipoMantencion.query.order_by(EquipoMantencion.id.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    headers = ['ID', 'Tipo', 'Marca', 'Modelo', 'N/S', 'Estado', 'Asignado A']
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        
    for eq in equipos:
        ws.append([
            eq.id,
            eq.tipo_equipo,
            eq.marca,
            eq.modelo,
            eq.numero_serie,
            eq.estado,
            eq.usuario_asignado.nombre if eq.usuario_asignado else 'Sin Asignar'
        ])
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, download_name="inventario_reporte.xlsx", as_attachment=True)

