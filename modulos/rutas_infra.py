from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, g, send_file, Response, session
from datetime import datetime, timedelta, date
from extensions import db
from models import *
from utils import login_required, requiere_permiso
import os
import json
from sqlalchemy import func
from werkzeug.security import generate_password_hash, check_password_hash


infra_bp = Blueprint('infra', __name__, template_folder='templates')

# ==================== GESTIÓN DE PUERTOS ====================

@infra_bp.route('/puertos')
@login_required
def puertos():
    if g.usuario.rol != 'admin':
        flash('Acceso denegado. Solo administradores.', 'error')
        return redirect(url_for('tickets.panel_agente'))
    search_query = request.args.get('q', '').strip()
    query = Puerto.query
    if search_query:
        query = query.filter(
            Puerto.nombre_servicio.ilike(f'%{search_query}%') |
            Puerto.numeros_puerto.ilike(f'%{search_query}%') |
            Puerto.descripcion.ilike(f'%{search_query}%')
        )
    puertos_list = query.order_by(Puerto.nombre_servicio.asc()).all()
    return render_template('puertos/lista.html', puertos=puertos_list, q=search_query)

@infra_bp.route('/puertos/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_puerto():
    if g.usuario.rol != 'admin':
        flash('Acceso denegado. Solo administradores.', 'error')
        return redirect(url_for('tickets.panel_agente'))
    if request.method == 'POST':
        try:
            nuevo_p = Puerto(
                nombre_servicio=request.form.get('nombre_servicio'),
                numeros_puerto=request.form.get('numeros_puerto'),
                descripcion=request.form.get('descripcion')
            )
            db.session.add(nuevo_p)
            db.session.commit()
            flash('Puerto registrado exitosamente', 'success')
            return redirect(url_for('infra.puertos'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al registrar puerto: {str(e)}', 'error')
    
    return render_template('puertos/formulario.html', puerto=None)

@infra_bp.route('/puertos/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_puerto(id):
    if g.usuario.rol != 'admin':
        flash('Acceso denegado. Solo administradores.', 'error')
        return redirect(url_for('tickets.panel_agente'))
    puerto = Puerto.query.get_or_404(id)
    if request.method == 'POST':
        try:
            puerto.nombre_servicio = request.form.get('nombre_servicio')
            puerto.numeros_puerto = request.form.get('numeros_puerto')
            puerto.descripcion = request.form.get('descripcion')
            db.session.commit()
            flash('Puerto actualizado exitosamente', 'success')
            return redirect(url_for('infra.puertos'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar puerto: {str(e)}', 'error')
            
    return render_template('puertos/formulario.html', puerto=puerto)

@infra_bp.route('/puertos/eliminar/<int:id>', methods=['POST'])
@login_required
def eliminar_puerto(id):
    if g.usuario.rol != 'admin':
        flash('Acceso denegado. Solo administradores.', 'error')
        return redirect(url_for('tickets.panel_agente'))
    puerto = Puerto.query.get_or_404(id)
    try:
        db.session.delete(puerto)
        db.session.commit()
        flash('Puerto eliminado exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar puerto: {str(e)}', 'error')
    return redirect(url_for('infra.puertos'))

# ==================== GESTIÓN DE LICENCIAS ====================

@infra_bp.route('/licencias')
@login_required
@requiere_permiso('ver_licencias')
def licencias():
    
    import datetime as _dt
    hoy = _dt.date.today()
    
    tab = request.args.get('tab', 'todas').strip()
    search_query = request.args.get('q', '').strip()
    
    query = Licencia.query
    
    # Filtro por tipo (tab)
    if tab == 'ssl':
        query = query.filter(Licencia.tipo == 'SSL')
    elif tab == 'software':
        query = query.filter(Licencia.tipo == 'Software')
    elif tab == 'saas':
        query = query.filter(Licencia.tipo == 'SaaS')
    
    # Búsqueda
    if search_query:
        query = query.filter(
            Licencia.nombre_servicio.ilike(f'%{search_query}%') |
            Licencia.proveedor.ilike(f'%{search_query}%') |
            Licencia.responsable.ilike(f'%{search_query}%')
        )
    
    page = request.args.get('page', 1, type=int)
    paginado = query.order_by(Licencia.fecha_expiracion.asc()).paginate(
        page=page, per_page=15, error_out=False)
    licencias_list = paginado.items
    
    # Estadísticas globales (sin filtro de tab/búsqueda)
    todas = Licencia.query.all()
    total = len(todas)
    total_ssl = sum(1 for l in todas if l.tipo == 'SSL')
    total_software = sum(1 for l in todas if l.tipo == 'Software')
    total_saas = sum(1 for l in todas if l.tipo == 'SaaS')
    expiradas = sum(1 for l in todas if l.fecha_expiracion and l.fecha_expiracion < hoy)
    proximas = sum(1 for l in todas if l.fecha_expiracion and 0 <= (l.fecha_expiracion - hoy).days <= 30)
    vigentes = sum(1 for l in todas if l.fecha_expiracion and (l.fecha_expiracion - hoy).days > 30)
    
    stats = {
        'total': total,
        'total_ssl': total_ssl,
        'total_software': total_software,
        'total_saas': total_saas,
        'expiradas': expiradas,
        'proximas': proximas,
        'vigentes': vigentes
    }
    
    licencias_sin_renovacion = [
        l for l in todas 
        if not l.renovacion_automatica and l.fecha_expiracion and (l.fecha_expiracion - hoy).days <= 30
    ]
    
    return render_template('licencias.html',
        licencias=licencias_list,
        paginado=paginado,
        vista='licencias',
        stats=stats,
        current_tab=tab,
        search_query=search_query,
        licencias_alerta_manual=licencias_sin_renovacion
    )

@infra_bp.route('/licencias/nueva', methods=['POST'])
@login_required
@requiere_permiso('ver_licencias')
def nueva_licencia():
    if g.usuario.rol not in ['admin', 'agente']:
        return jsonify({'success': False, 'message': 'Acceso denegado'})
    
    import datetime as _dt
    try:
        f_inicio = request.form.get('fecha_inicio')
        f_expiracion = request.form.get('fecha_expiracion')
        
        def parse_date(d_str):
            if not d_str:
                return None
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y'):
                try:
                    return _dt.datetime.strptime(d_str.strip(), fmt).date()
                except ValueError:
                    pass
            raise ValueError(f"Formato de fecha inválido: {d_str}")
        
        f_inicio_obj = parse_date(f_inicio)
        f_exp_obj = parse_date(f_expiracion)
        
        if not f_exp_obj:
            flash('La fecha de expiración es obligatoria.', 'error')
            return redirect(url_for('infra.licencias'))

        lic = Licencia(
            nombre_servicio=request.form.get('nombre_servicio'),
            tipo=request.form.get('tipo'),
            proveedor=request.form.get('proveedor'),
            cantidad=request.form.get('cantidad') or None,
            responsable=request.form.get('responsable'),
            fecha_inicio=f_inicio_obj,
            fecha_expiracion=f_exp_obj,
            renovacion_automatica=request.form.get('renovacion_automatica') == 'on',
            estado=request.form.get('estado', 'Activo'),
            observaciones=request.form.get('observaciones')
        )
        db.session.add(lic)
        db.session.commit()
        flash('Licencia agregada exitosamente.', 'success')
    except ValueError:
        flash('Error al agregar licencia: El formato de la fecha es incorrecto.', 'error')
    except Exception as e:
        msg = str(e).replace("'", "")
        flash(f'Error al agregar licencia: {msg}', 'error')
        
    return redirect(url_for('infra.licencias'))

@infra_bp.route('/licencias/editar/<int:id>', methods=['POST'])
@login_required
@requiere_permiso('ver_licencias')
def editar_licencia(id):
    if g.usuario.rol not in ['admin', 'agente']:
        return jsonify({'success': False, 'message': 'Acceso denegado'})
    
    lic = db.session.get(Licencia, id)
    if not lic:
        flash('Licencia no encontrada.', 'error')
        return redirect(url_for('infra.licencias'))
        
    import datetime as _dt
    try:
        f_inicio = request.form.get('fecha_inicio')
        f_expiracion = request.form.get('fecha_expiracion')
        
        def parse_date(d_str):
            if not d_str:
                return None
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y'):
                try:
                    return _dt.datetime.strptime(d_str.strip(), fmt).date()
                except ValueError:
                    pass
            raise ValueError(f"Formato de fecha inválido: {d_str}")
        
        f_inicio_obj = parse_date(f_inicio)
        f_exp_obj = parse_date(f_expiracion)
        
        if not f_exp_obj:
            flash('La fecha de expiración es obligatoria.', 'error')
            return redirect(url_for('infra.licencias'))

        lic.nombre_servicio = request.form.get('nombre_servicio')
        lic.tipo = request.form.get('tipo')
        lic.proveedor = request.form.get('proveedor')
        lic.cantidad = request.form.get('cantidad') or None
        lic.responsable = request.form.get('responsable')
        lic.fecha_inicio = f_inicio_obj
        lic.fecha_expiracion = f_exp_obj
        lic.renovacion_automatica = (request.form.get('renovacion_automatica') == 'on')
        lic.estado = request.form.get('estado', 'Activo')
        lic.observaciones = request.form.get('observaciones')

        db.session.commit()
        flash('Licencia actualizada exitosamente.', 'success')
    except ValueError:
        flash('Error al actualizar licencia: El formato de la fecha es incorrecto.', 'error')
    except Exception as e:
        msg = str(e).replace("'", "")
        flash(f'Error al actualizar licencia: {msg}', 'error')
        
    return redirect(url_for('infra.licencias'))

@infra_bp.route('/licencias/eliminar/<int:id>', methods=['POST'])
@login_required
@requiere_permiso('ver_licencias')
def eliminar_licencia(id):
    if g.usuario.rol not in ['admin', 'agente']:
        return jsonify({'success': False, 'message': 'Acceso denegado'})
    lic = db.session.get(Licencia, id)
    if lic:
        db.session.delete(lic)
        db.session.commit()
        flash('Licencia eliminada.', 'success')
    return redirect(url_for('infra.licencias'))

# ---------- EVENTOS Y ACCIONES CORRECTIVAS ----------

@infra_bp.route('/eventos/dashboard')
@login_required
def dashboard_eventos():
    """Redirige al dashboard de hallazgos (módulo actual de eventos)."""
    return redirect(url_for('hallazgos.dashboard'))

@infra_bp.route('/eventos')
@login_required
def eventos_lista():
    """Lista de eventos registrados"""
    eventos = Evento.query.order_by(Evento.fecha_registro.desc()).all()
    return render_template('eventos/lista_eventos.html', eventos=eventos)

@infra_bp.route('/eventos_old/nuevo', methods=['GET', 'POST'])
@login_required
def eventos_nuevo():
    """Registrar nuevo evento"""
    if request.method == 'POST':
        area_id = request.form.get('area_id')
        tipo_evento_id = request.form.get('tipo_evento_id')
        descripcion = request.form.get('descripcion')
        
        # Generar código único para el evento
        from datetime import datetime
        year = datetime.now().year
        ultimo_evento = Evento.query.order_by(Evento.id.desc()).first()
        numero = (ultimo_evento.id + 1) if ultimo_evento else 1
        codigo = f"EV-{year}-{numero:04d}"

        nuevo_evento = Evento(
            codigo=codigo,
            area_id=area_id,
            tipo_evento_id=tipo_evento_id,
            descripcion=descripcion,
            responsable_id=g.usuario.id
        )
        try:
            db.session.add(nuevo_evento)
            db.session.commit()
            flash('Evento registrado.', 'success')
            return redirect(url_for('eventos_legacy.eventos_lista'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al registrar el evento: {str(e)}', 'error')
    
    areas = Area.query.filter_by(activa=True).all()
    tipos = TipoEvento.query.filter_by(activo=True).all()
    sistemas = SistemaNormativo.query.filter_by(activo=True).all()
    
    return render_template('eventos/formulario_evento.html', areas=areas, tipos=tipos, sistemas=sistemas)

@infra_bp.route('/eventos/<int:id>')
@login_required
def ver_evento(id):
    """Detalle de evento"""
    evento = db.session.get(Evento, id)
    return render_template('eventos/ver_evento.html', evento=evento)

@infra_bp.route('/acciones_correctivas')
@login_required
def acciones_lista():
    """Lista de Acciones Correctivas"""
    acciones = AccionCorrectiva.query.order_by(AccionCorrectiva.fecha_registro.desc()).all()
    return render_template('eventos/lista_acciones.html', acciones=acciones)

@infra_bp.route('/acciones_correctivas/<int:id>/acr')
@login_required
def acciones_acr(id):
    """Análisis de Causa Raíz (ACR) iterativo"""
    accion = db.session.get(AccionCorrectiva, id)
    return render_template('eventos/acr.html', accion=accion)

@infra_bp.route('/configuraciones/eventos', methods=['GET', 'POST'])
@login_required
def configuraciones_eventos():
    """Catálogos y configuraciones del sistema de eventos"""
    if g.usuario.rol != 'admin':
        flash('Acceso denegado', 'error')
        return redirect(url_for('tickets.index'))
    return render_template('eventos/configuraciones.html')

from io import BytesIO
from flask import send_file
from openpyxl.styles import Font, PatternFill

@infra_bp.route('/exportar/tickets')
@login_required
def exportar_tickets():
    if g.usuario.rol == 'cliente' and not g.usuario.tiene_permiso('ver_tickets_area'):
        flash('Acceso denegado', 'error')
        return redirect(url_for('tickets.mis_tickets'))
        
    query = Ticket.query.join(Usuario, Ticket.usuario_id == Usuario.id)
    if g.usuario.rol not in ['admin', 'agente']:
        if g.usuario.area_id:
            query = query.filter(Usuario.area_id == g.usuario.area_id)
        else:
            query = query.filter(Usuario.id == -1)
            
    tickets = query.order_by(Ticket.fecha_creacion.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"
    
    headers = ['ID', 'Asunto', 'Usuario', 'Estado', 'Prioridad', 'Fecha Creación', 'Técnico Asignado']
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        
    for t in tickets:
        tecnico_nombre = t.tecnico.nombre if t.tecnico else 'Sin asignar'
        ws.append([
            t.id, 
            t.asunto, 
            t.usuario.nombre if t.usuario else 'Desconocido', 
            t.estado, 
            t.prioridad, 
            t.fecha_creacion.strftime('%d/%m/%Y %H:%M') if t.fecha_creacion else '',
            tecnico_nombre
        ])
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, download_name="tickets_reporte.xlsx", as_attachment=True)

@infra_bp.route('/exportar/inventario')
@login_required
@requiere_permiso('ver_inventario')
def exportar_inventario():
    equipos = EquipoMantencion.query.order_by(EquipoMantencion.id.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    headers = ['ID', 'Código', 'Nombre', 'Marca', 'Modelo', 'Serie', 'Área', 'Responsable', 'Estado', 'Frec. Mantención', 'Próxima Mantención', 'Alerta']
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        
    for eq in equipos:
        ws.append([
            eq.id,
            eq.codigo or '',
            eq.nombre or '',
            eq.marca or '',
            eq.modelo or '',
            eq.serie or '',
            eq.area or '',
            eq.responsable or 'Sin Asignar',
            eq.estado or '',
            eq.frecuencia_mantencion or '',
            eq.proxima_mantencion or '',
            eq.estado_alerta
        ])
    
    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 3, 40)
    
    # === Hoja 2: Responsables con más de un equipo ===
    from collections import defaultdict
    responsable_equipos = defaultdict(list)
    for eq in equipos:
        resp = (eq.responsable or '').strip()
        if resp:
            responsable_equipos[resp].append(eq)
    
    # Filtrar solo los que tienen más de 1 equipo
    multi_responsables = {r: eqs for r, eqs in responsable_equipos.items() if len(eqs) > 1}
    
    ws2 = wb.create_sheet(title="Usuarios Múltiples Equipos")
    ws2_headers = ['Responsable', 'Cant. Equipos', 'ID Equipo', 'Código', 'Nombre Equipo', 'Marca', 'Modelo', 'Estado']
    ws2.append(ws2_headers)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    highlight_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    for resp_nombre in sorted(multi_responsables.keys()):
        eqs = multi_responsables[resp_nombre]
        for i, eq in enumerate(eqs):
            row = [
                resp_nombre if i == 0 else '',
                len(eqs) if i == 0 else '',
                eq.id,
                eq.codigo or '',
                eq.nombre or '',
                eq.marca or '',
                eq.modelo or '',
                eq.estado or ''
            ]
            ws2.append(row)
            # Resaltar la fila
            for cell in ws2[ws2.max_row]:
                cell.fill = highlight_fill
        # Fila vacía separadora entre responsables
        ws2.append([])
    
    for col in ws2.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws2.column_dimensions[col_letter].width = min(max_length + 3, 40)
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, download_name="inventario_reporte.xlsx", as_attachment=True)


