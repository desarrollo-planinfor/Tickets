from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, g, send_file, Response, session
from datetime import datetime, timedelta, date
from extensions import db

from sqlalchemy.orm import joinedload
from models import *
from utils import login_required, requiere_permiso, calcular_minutos_habiles, parsear_fecha_hora
from cron_jobs import email_ticket_recibido
import os
import json
from sqlalchemy import func
from werkzeug.security import generate_password_hash, check_password_hash


tickets_bp = Blueprint('tickets', __name__, template_folder='templates')

# ==================== RUTAS DE LA APLICACIÃ“N ====================

@tickets_bp.route('/')
def index():
    """PÃ¡gina principal - Redirecciona segÃºn rol"""
    if 'usuario_id' in session:
        usuario = db.session.get(Usuario, session['usuario_id'])
        if usuario and usuario.rol == 'admin':
            return redirect(url_for('admin.panel_admin'))
        elif usuario and usuario.rol == 'agente':
            return redirect(url_for('tickets.panel_agente'))
        elif usuario and usuario.rol == 'cliente':
            return redirect(url_for('tickets.mis_tickets'))
    return redirect(url_for('auth.login'))

# ---------- AUTENTICACIÃ“N ----------

# ---------- PORTAL CLIENTE ----------

@tickets_bp.route('/portal')
@login_required
def portal_cliente():
    """Portal del cliente para crear tickets"""
    return render_template('portal_cliente.html')

@tickets_bp.route('/portal/crear', methods=['POST'])
@login_required
def crear_ticket():
    """Crear nuevo ticket desde portal cliente"""
    try:
        usuario = g.usuario
        asunto = request.form.get('asunto')
        descripcion = request.form.get('descripcion')
        categoria = request.form.get('categoria', 'general')
        prioridad = 'Media'  # Solo modificable por admin, default Media
        departamento = 'soporte' # Default soporte
        
        ticket = Ticket(
            usuario_id=usuario.id,
            asunto=asunto,
            descripcion=descripcion,
            categoria=categoria,
            prioridad=prioridad,
            departamento=departamento,
            estado='PENDIENTE'
        )
        db.session.add(ticket)
        
        # Crear log inicial
        log = TicketLog(
            ticket_id=None,  # Se actualiza despuÃ©s
            estado_nuevo='PENDIENTE',
            descripcion='Ticket creado por el cliente',
            usuario_id=usuario.id
        )
        
        # Commit para obtener IDs
        db.session.flush()
        log.ticket_id = ticket.id
        db.session.commit()
        
        # Crear notificaciÃ³n de confirmaciÃ³n al cliente (eliminada a peticiÃ³n)
        # Solo dejaremos la alerta a soporte TI
        
        # Alerta para el equipo de soporte TI
        alerta_ti = Notificacion(
            ticket_id=ticket.id,
            destinatario='soporte.ti@planinfor.cl',
            asunto=f'[Soporte TI] Nuevo Ticket #{ticket.id} - {asunto}',
            mensaje=email_nuevo_ticket(ticket.id, asunto, descripcion, categoria, usuario.nombre, usuario.email),
            tipo='email'
        )
        db.session.add(alerta_ti)
        
        db.session.commit()
        
        flash(f'Ticket #{ticket.id} creado exitosamente. Te enviaremos una notificaciÃ³n cuando sea atendido.', 'success')
        return redirect(url_for('tickets.portal_cliente'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al crear ticket: {str(e)}', 'error')
        return redirect(url_for('tickets.portal_cliente'))


@tickets_bp.route('/portal/mis-tickets')
@login_required
def mis_tickets():
    """Ver tickets del cliente (requiere login)"""
    usuario = g.usuario
    tickets = Ticket.query.options(joinedload(Ticket.tecnico)).filter_by(usuario_id=usuario.id).order_by(Ticket.fecha_creacion.desc()).all()
    return render_template('mis_tickets.html', tickets=tickets, usuario=usuario)

@tickets_bp.route('/mis-tickets')
@login_required
def mis_tickets_alt():
    """Ver tickets del cliente - ruta alternativa"""
    return redirect(url_for('tickets.mis_tickets'))

# ---------- PANEL DE AGENTE ----------

@tickets_bp.route('/agente')
@login_required
def panel_agente():
    """Dashboard del agente con semÃ¡foro de prioridades"""
    # Tickets pendientes sin hora de atenciÃ³n o atrasados (ROJO)
    tickets_rojos = Ticket.query.options(joinedload(Ticket.usuario), joinedload(Ticket.tecnico)).filter(
        Ticket.estado.in_(['PENDIENTE', 'ATRASADO'])
    ).order_by(Ticket.fecha_creacion.asc()).all()
    
    # Tickets recibidos con hora de atenciÃ³n programada (AMARILLO)
    tickets_amarillos = Ticket.query.options(joinedload(Ticket.usuario), joinedload(Ticket.tecnico)).filter(
        Ticket.estado == 'RECIBIDO'
    ).order_by(Ticket.fecha_atencion_programada.asc()).all()
    
    # Tickets en proceso (VERDE)
    tickets_verdes = Ticket.query.options(joinedload(Ticket.usuario), joinedload(Ticket.tecnico)).filter(
        Ticket.estado == 'EN_PROCESO'
    ).order_by(Ticket.fecha_inicio_atencion.asc()).all()
    
    # Tickets cerrados recientemente
    tickets_cerrados = Ticket.query.options(joinedload(Ticket.usuario), joinedload(Ticket.tecnico)).filter(
        Ticket.estado.in_(['RESUELTO', 'CERRADO'])
    ).order_by(Ticket.fecha_cierre.desc()).limit(10).all()
    
    # EstadÃ­sticas
    stats = {
        'total': Ticket.query.count(),
        'pendientes': Ticket.query.options(joinedload(Ticket.usuario), joinedload(Ticket.tecnico)).filter(Ticket.estado.in_(['PENDIENTE', 'ATRASADO'])).count(),
        'en_proceso': Ticket.query.options(joinedload(Ticket.usuario), joinedload(Ticket.tecnico)).filter(Ticket.estado.in_(['RECIBIDO', 'EN_PROCESO'])).count(),
        'cerrados': Ticket.query.options(joinedload(Ticket.usuario), joinedload(Ticket.tecnico)).filter(Ticket.estado.in_(['RESUELTO', 'CERRADO'])).count()
    }
    
    return render_template('panel_agente.html', 
                         tickets_rojos=tickets_rojos,
                         tickets_amarillos=tickets_amarillos,
                         tickets_verdes=tickets_verdes,
                         tickets_cerrados=tickets_cerrados,
                         stats=stats)

@tickets_bp.route('/agente/recibir/<int:ticket_id>', methods=['POST'])
@login_required
def recibir_ticket(ticket_id):
    """Admin/Agente marca ticket como recibido y asigna hora de atenciÃ³n"""
    try:
        hora_atencion = request.form.get('hora_atencion')
        nueva_prioridad = request.form.get('prioridad')
        nuevo_departamento = request.form.get('departamento')
        
        ticket = db.session.get(Ticket, ticket_id)
        if not ticket:
            flash('Ticket no encontrado', 'error')
            return redirect(url_for('tickets.panel_agente'))
        
        ticket.estado = 'RECIBIDO'
        ticket.fecha_recepcion = datetime.now()
        
        if hora_atencion:
            ticket.fecha_atencion_programada = parsear_fecha_hora(hora_atencion)
            
        if nueva_prioridad:
            ticket.prioridad = nueva_prioridad
        if nuevo_departamento:
            ticket.departamento = nuevo_departamento
        
        # Log de cambio
        log = TicketLog(
            ticket_id=ticket.id,
            estado_anterior='PENDIENTE',
            estado_nuevo='RECIBIDO',
            descripcion=f'Ticket recibido. Hora de atenciÃ³n programada: {hora_atencion}'
        )
        db.session.add(log)
        db.session.commit()
        
        # Notificar al equipo de TI
        nombre_usuario = ticket.usuario.nombre if ticket.usuario else 'Desconocido'
        notificacion = Notificacion(
            ticket_id=ticket.id,
            destinatario='soporte.ti@planinfor.cl',
            asunto=f'[Soporte TI] Ticket #{ticket.id} Aceptado - AtenciÃ³n Programada',
            mensaje=email_ticket_recibido(ticket.id, ticket.asunto, nombre_usuario, hora_atencion),
            tipo='email'
        )
        db.session.add(notificacion)
        db.session.commit()
        
        flash(f'Ticket #{ticket_id} marcado como recibido', 'success')
        return redirect(url_for('tickets.panel_agente'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('tickets.panel_agente'))

@tickets_bp.route('/agente/atender/<int:ticket_id>')
@login_required
def atender_ticket(ticket_id):
    """TÃ©cnico inicia atenciÃ³n del ticket"""
    try:
        ticket = db.session.get(Ticket, ticket_id)
        if not ticket:
            flash('Ticket no encontrado', 'error')
            return redirect(url_for('tickets.panel_agente'))
        
        ticket.estado = 'EN_PROCESO'
        ticket.fecha_inicio_atencion = datetime.now()
        ticket.tecnico_id = g.usuario.id  # Asignar tÃ©cnico que atiende
        
        # Calcular tiempo de respuesta
        if ticket.fecha_creacion:
            ticket.tiempo_respuesta = calcular_minutos_habiles(ticket.fecha_creacion, ticket.fecha_inicio_atencion)
        
        log = TicketLog(
            ticket_id=ticket.id,
            usuario_id=g.usuario.id,
            estado_anterior='RECIBIDO',
            estado_nuevo='EN_PROCESO',
            descripcion=f'AtenciÃ³n iniciada por {g.usuario.nombre}'
        )
        db.session.add(log)
        db.session.commit()
        
        flash(f'Ticket #{ticket_id} en proceso', 'success')
        return redirect(url_for('tickets.panel_agente'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('tickets.panel_agente'))

@tickets_bp.route('/agente/cerrar/<int:ticket_id>', methods=['POST'])
@login_required
def cerrar_ticket(ticket_id):
    """Cerrar ticket (inmediato o con estimaciÃ³n)"""
    try:
        tipo_cierre = request.form.get('tipo_cierre')  # inmediato, estimado
        notas = request.form.get('notas', '')
        
        ticket = db.session.get(Ticket, ticket_id)
        if not ticket:
            flash('Ticket no encontrado', 'error')
            return redirect(url_for('tickets.panel_agente'))
        
        ticket.fecha_cierre = datetime.now()
        ticket.notas = notas
        
        if tipo_cierre == 'inmediato':
            ticket.estado = 'CERRADO'
            estado_final = 'CERRADO'
        else:
            # Con estimaciÃ³n - queda en proceso
            tiempo_estimado = request.form.get('tiempo_estimado', 60)
            ticket.tiempo_estimado = int(tiempo_estimado)
            ticket.estado = 'EN_PROCESO'
            estado_final = 'EN_PROCESO (Con estimaciÃ³n)'
        
        # Calcular tiempo de resoluciÃ³n
        if ticket.fecha_inicio_atencion:
            ticket.tiempo_resolucion = calcular_minutos_habiles(ticket.fecha_inicio_atencion, ticket.fecha_cierre)
        
        log = TicketLog(
            ticket_id=ticket.id,
            estado_anterior='EN_PROCESO',
            estado_nuevo=estado_final,
            descripcion=f'Ticket cerrado. Notas: {notas}'
        )
        db.session.add(log)
        db.session.commit()
        
        flash(f'Ticket #{ticket_id} cerrado', 'success')
        return redirect(url_for('tickets.panel_agente'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('tickets.panel_agente'))

import uuid
from werkzeug.utils import secure_filename

@tickets_bp.route('/ticket/<int:ticket_id>/adjuntar', methods=['POST'])
@login_required
def adjuntar_archivo(ticket_id):
    ticket = db.session.get(Ticket, ticket_id)
    if not ticket:
        flash('Ticket no encontrado', 'error')
        return redirect(url_for('tickets.mis_tickets'))
        
    if g.usuario.rol == 'cliente' and ticket.usuario_id != g.usuario.id:
        flash('No tienes autorizaciÃ³n', 'error')
        return redirect(url_for('tickets.mis_tickets'))
        
    if 'archivo' not in request.files:
        flash('No se seleccionÃ³ ningÃºn archivo', 'error')
        return redirect(url_for('tickets.ver_ticket', ticket_id=ticket.id))
        
    archivo = request.files['archivo']
    if archivo.filename == '':
        flash('No se seleccionÃ³ ningÃºn archivo', 'error')
        return redirect(url_for('tickets.ver_ticket', ticket_id=ticket.id))
        
    if archivo:
        filename = secure_filename(archivo.filename)
        # Generate unique name
        unique_filename = f"{uuid.uuid4().hex}_{filename}"
        upload_folder = os.path.join(app.root_path, 'static', 'uploads', 'adjuntos')
        os.makedirs(upload_folder, exist_ok=True)
        
        file_path = os.path.join(upload_folder, unique_filename)
        archivo.save(file_path)
        
        adjunto = TicketAdjunto(
            ticket_id=ticket.id,
            usuario_id=g.usuario.id,
            nombre_archivo=filename,
            ruta_archivo=f"uploads/adjuntos/{unique_filename}"
        )
        db.session.add(adjunto)
        
        # Agregar log
        log = TicketLog(
            ticket_id=ticket.id,
            estado_anterior=ticket.estado,
            estado_nuevo=ticket.estado,
            descripcion=f"Archivo adjuntado: {filename}",
            usuario_id=g.usuario.id
        )
        db.session.add(log)
        
        db.session.commit()
        flash('Archivo adjuntado exitosamente', 'success')
        
    return redirect(url_for('tickets.ver_ticket', ticket_id=ticket.id))

@tickets_bp.route('/ticket/<int:ticket_id>')
@login_required
def ver_ticket(ticket_id):
    """Ver detalles de un ticket (Universal para cliente, agente y admin)"""
    ticket = Ticket.query.get_or_404(ticket_id)
    
    # Seguridad: Si es cliente, solo puede ver sus propios tickets
    if g.usuario.rol == 'cliente' and ticket.usuario_id != g.usuario.id:
        flash('No tienes autorizaciÃ³n para ver este ticket.', 'error')
        return redirect(url_for('tickets.mis_tickets'))
    
    # Obtener historial de cambios
    logs = TicketLog.query.filter_by(ticket_id=ticket_id).order_by(TicketLog.fecha_cambio.asc()).all()
    
    return render_template('ver_ticket.html', ticket=ticket, logs=logs)

# ---------- GESTIÃ“N DE ESTADOS DE TICKET ----------

@tickets_bp.route('/agente/aceptar/<int:ticket_id>')
@login_required
def aceptar_ticket(ticket_id):
    """Aceptar un ticket pendiente"""
    ticket = db.session.get(Ticket, ticket_id)
    if not ticket:
        flash('Ticket no encontrado', 'error')
        return redirect(url_for('tickets.panel_agente'))
    
    try:
        # Cambiar estado a RECIBIDO
        estado_anterior = ticket.estado
        ticket.estado = 'RECIBIDO'
        ticket.fecha_recepcion = datetime.now()
        
        # Calcular tiempo de respuesta
        if ticket.fecha_creacion:
            ticket.tiempo_respuesta = calcular_minutos_habiles(ticket.fecha_creacion, ticket.fecha_recepcion)
        
        # Registrar log
        log = TicketLog(
            ticket_id=ticket.id,
            usuario_id=g.usuario.id,
            estado_nuevo='RECIBIDO',
            descripcion=f'Ticket aceptado por {g.usuario.nombre}'
        )
        db.session.add(log)
        db.session.commit()
        
        flash(f'Ticket #{ticket_id} aceptado', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('tickets.ver_ticket', ticket_id=ticket_id))

@tickets_bp.route('/agente/iniciar/<int:ticket_id>')
@login_required
def iniciar_ticket(ticket_id):
    """Iniciar trabajo en un ticket"""
    ticket = db.session.get(Ticket, ticket_id)
    if not ticket:
        flash('Ticket no encontrado', 'error')
        return redirect(url_for('tickets.panel_agente'))
    
    try:
        ticket.estado = 'EN_PROCESO'
        ticket.fecha_inicio_atencion = datetime.now()
        ticket.tecnico_id = g.usuario.id
        
        # Registrar log
        log = TicketLog(
            ticket_id=ticket.id,
            usuario_id=g.usuario.id,
            estado_nuevo='EN_PROCESO',
            descripcion=f'AtenciÃ³n iniciada por {g.usuario.nombre}'
        )
        db.session.add(log)
        db.session.commit()
        
        flash(f'Ticket #{ticket_id} en proceso', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('tickets.ver_ticket', ticket_id=ticket_id))

@tickets_bp.route('/agente/pausar/<int:ticket_id>')
@login_required
def pausar_ticket(ticket_id):
    """Pausar un ticket en proceso"""
    ticket = db.session.get(Ticket, ticket_id)
    if not ticket:
        flash('Ticket no encontrado', 'error')
        return redirect(url_for('tickets.panel_agente'))
    
    try:
        ticket.estado = 'PENDIENTE'
        
        # Registrar log
        log = TicketLog(
            ticket_id=ticket.id,
            usuario_id=g.usuario.id,
            estado_nuevo='PENDIENTE',
            descripcion=f'Ticket pausado por {g.usuario.nombre}'
        )
        db.session.add(log)
        db.session.commit()
        
        flash(f'Ticket #{ticket_id} pausado', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('tickets.ver_ticket', ticket_id=ticket_id))

@tickets_bp.route('/agente/reabrir/<int:ticket_id>')
@login_required
def reabrir_ticket(ticket_id):
    """Reabrir un ticket cerrado"""
    ticket = db.session.get(Ticket, ticket_id)
    if not ticket:
        flash('Ticket no encontrado', 'error')
        return redirect(url_for('tickets.panel_agente'))
    
    try:
        ticket.estado = 'PENDIENTE'
        ticket.fecha_cierre = None
        ticket.notas = None
        
        # Registrar log
        log = TicketLog(
            ticket_id=ticket.id,
            usuario_id=g.usuario.id,
            estado_nuevo='PENDIENTE',
            descripcion=f'Ticket reopen by {g.usuario.nombre}'
        )
        db.session.add(log)
        db.session.commit()
        
        flash(f'Ticket #{ticket_id} reopen', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('tickets.ver_ticket', ticket_id=ticket_id))

# ---------- PANEL DE ADMINISTRADOR ----------

# ==================== API ====================

@tickets_bp.route('/api/tickets', methods=['GET'])
def api_get_tickets():
    """API para obtener todos los tickets con relaciones, omitiendo contraseÃ±as"""
    include_logs = request.args.get('logs', 'true').lower() != 'false'
    tickets = Ticket.query.options(joinedload(Ticket.usuario), joinedload(Ticket.tecnico)).order_by(Ticket.fecha_creacion.desc()).all()
    tickets_list = []
    
    for t in tickets:
        # Serializar usuario creador
        creador = None
        if t.usuario:
            creador = {
                'id': t.usuario.id,
                'nombre': t.usuario.nombre,
                'email': t.usuario.email,
                'rol': t.usuario.rol,
                'activo': t.usuario.activo,
                'created_at': t.usuario.created_at.isoformat() if t.usuario.created_at else None
            }
            
        # Serializar tÃ©cnico asignado
        tecnico = None
        if t.tecnico:
            tecnico = {
                'id': t.tecnico.id,
                'nombre': t.tecnico.nombre,
                'email': t.tecnico.email,
                'rol': t.tecnico.rol,
                'activo': t.tecnico.activo
            }
            
        # Serializar logs
        logs_list = []
        if include_logs:
            for log in t.logs:
                log_usuario = None
                if log.usuario:
                    log_usuario = {
                        'id': log.usuario.id,
                        'nombre': log.usuario.nombre,
                        'email': log.usuario.email,
                        'rol': log.usuario.rol
                    }
                logs_list.append({
                    'id': log.id,
                    'estado_anterior': log.estado_anterior,
                    'estado_nuevo': log.estado_nuevo,
                    'descripcion': log.descripcion,
                    'fecha_cambio': log.fecha_cambio.isoformat() if log.fecha_cambio else None,
                    'usuario': log_usuario
                })
            
        ticket_data = {
            'id': t.id,
            'asunto': t.asunto,
            'descripcion': t.descripcion,
            'categoria': t.categoria,
            'prioridad': t.prioridad,
            'departamento': t.departamento,
            'estado': t.estado,
            'fecha_creacion': t.fecha_creacion.isoformat() if t.fecha_creacion else None,
            'fecha_recepcion': t.fecha_recepcion.isoformat() if t.fecha_recepcion else None,
            'fecha_atencion_programada': t.fecha_atencion_programada.isoformat() if t.fecha_atencion_programada else None,
            'fecha_inicio_atencion': t.fecha_inicio_atencion.isoformat() if t.fecha_inicio_atencion else None,
            'fecha_cierre': t.fecha_cierre.isoformat() if t.fecha_cierre else None,
            'tiempo_respuesta': t.tiempo_respuesta,
            'tiempo_resolucion': t.tiempo_resolucion,
            'tiempo_estimado': t.tiempo_estimado,
            'notas': t.notas,
            'creador': creador,
            'tecnico': tecnico
        }
        
        if include_logs:
            ticket_data['historial_logs'] = logs_list
            
        tickets_list.append(ticket_data)
        
    return jsonify({'status': 'success', 'total': len(tickets_list), 'data': tickets_list})

@tickets_bp.route('/api/tickets-simple', methods=['GET'])
def api_get_tickets_simple():
    """API para obtener todos los tickets con relaciones, omitiendo contraseÃ±as e historial de logs"""
    tickets = Ticket.query.options(joinedload(Ticket.usuario), joinedload(Ticket.tecnico)).order_by(Ticket.fecha_creacion.desc()).all()
    tickets_list = []
    
    for t in tickets:
        # Serializar usuario creador
        creador = None
        if t.usuario:
            creador = {
                'id': t.usuario.id,
                'nombre': t.usuario.nombre,
                'email': t.usuario.email,
                'rol': t.usuario.rol,
                'activo': t.usuario.activo,
                'created_at': t.usuario.created_at.isoformat() if t.usuario.created_at else None
            }
            
        # Serializar tÃ©cnico asignado
        tecnico = None
        if t.tecnico:
            tecnico = {
                'id': t.tecnico.id,
                'nombre': t.tecnico.nombre,
                'email': t.tecnico.email,
                'rol': t.tecnico.rol,
                'activo': t.tecnico.activo
            }
            
        ticket_data = {
            'id': t.id,
            'asunto': t.asunto,
            'descripcion': t.descripcion,
            'categoria': t.categoria,
            'prioridad': t.prioridad,
            'departamento': t.departamento,
            'estado': t.estado,
            'fecha_creacion': t.fecha_creacion.isoformat() if t.fecha_creacion else None,
            'fecha_recepcion': t.fecha_recepcion.isoformat() if t.fecha_recepcion else None,
            'fecha_atencion_programada': t.fecha_atencion_programada.isoformat() if t.fecha_atencion_programada else None,
            'fecha_inicio_atencion': t.fecha_inicio_atencion.isoformat() if t.fecha_inicio_atencion else None,
            'fecha_cierre': t.fecha_cierre.isoformat() if t.fecha_cierre else None,
            'tiempo_respuesta': t.tiempo_respuesta,
            'tiempo_resolucion': t.tiempo_resolucion,
            'tiempo_estimado': t.tiempo_estimado,
            'notas': t.notas,
            'creador': creador,
            'tecnico': tecnico
        }
        
        tickets_list.append(ticket_data)
        
    return jsonify({'status': 'success', 'total': len(tickets_list), 'data': tickets_list})

@tickets_bp.route('/todos')
@login_required
def todos_tickets():
    """Ver todos los tickets con paginación (Agentes, Admin, o permisos de área)"""
    if g.usuario.rol == 'cliente' and not g.usuario.tiene_permiso('ver_tickets_area'):
        flash('Acceso denegado.', 'error')
        return redirect(url_for('tickets.mis_tickets'))
        
    page = request.args.get('page', 1, type=int)
    search_query = request.args.get('q', '').strip()
    
    query = Ticket.query.join(Usuario, Ticket.usuario_id == Usuario.id)
    
    # Filtro de área si no es admin ni agente (es un usuario con permiso de área)
    if g.usuario.rol not in ['admin', 'agente']:
        if g.usuario.area_id:
            query = query.filter(Usuario.area_id == g.usuario.area_id)
        else:
            query = query.filter(Usuario.id == -1) # No tiene área, no ve tickets de otros
    
    if search_query:
        query = query.filter(Usuario.nombre.ilike(f'%{search_query}%'))
        
    tickets_paginados = query.order_by(Ticket.fecha_creacion.desc()).paginate(page=page, per_page=15, error_out=False)
    
    return render_template('todos_tickets.html', tickets_paginados=tickets_paginados, search_query=search_query)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

@tickets_bp.route('/exportar')
@login_required
def exportar_tickets():
    if g.usuario.rol == 'cliente' and not g.usuario.tiene_permiso('ver_tickets_area'):
        flash('Acceso denegado', 'error')
        return redirect(url_for('tickets.mis_tickets'))
        
    query = Ticket.query.join(Usuario, Ticket.usuario_id == Usuario.id)
    if g.usuario.rol not in ['admin', 'agente']:
        if g.usuario.area_id:
            query = query.filter(Usuario.area_id == g.usuario.area_id)
        else:
            query = query.filter(Usuario.id == -1)
            
    tickets = query.order_by(Ticket.fecha_creacion.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"
    
    headers = ['ID', 'Asunto', 'Usuario', 'Estado', 'Prioridad', 'Fecha Creacion', 'Tecnico Asignado']
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(vertical='center', wrap_text=True)
    thin = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin
        
    for t in tickets:
        tecnico_nombre = t.tecnico.nombre if t.tecnico else 'Sin asignar'
        ws.append([
            t.id, 
            t.asunto, 
            t.usuario.nombre if t.usuario else 'Desconocido', 
            (t.estado or '').replace('_', ' '), 
            t.prioridad, 
            t.fecha_creacion.strftime('%d/%m/%Y %H:%M') if t.fecha_creacion else '',
            tecnico_nombre
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = cell_align
            cell.border = thin

    # Anchos de columna según contenido (evita texto cortado/superpuesto)
    min_widths = {
        1: 8,   # ID
        2: 28,  # Asunto
        3: 18,  # Usuario
        4: 14,  # Estado
        5: 12,  # Prioridad
        6: 18,  # Fecha
        7: 20,  # Tecnico
    }
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for cell in ws[get_column_letter(col_idx)]:
            val = '' if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        width = max(min_widths.get(col_idx, 12), min(max_len + 2, 45))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, download_name="tickets_reporte.xlsx", as_attachment=True)

